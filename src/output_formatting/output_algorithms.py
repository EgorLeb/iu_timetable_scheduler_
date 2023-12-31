from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import os

WEEKDAYS = ('Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun')
PERIODS = ("9:00-10:30", "10:40-12:10", "12:40-14:10", "14:20-15:50", "16:00-17:30", "17:40-19:00")


def parametrized(week):
    result = dict()
    # format: (weekday, class_number, group_name) -> (room_number, class_name, instructor)
    for weekday in WEEKDAYS:
        for class_number in range(len(week[weekday])):
            for room_number in week[weekday][class_number].keys():
                slot = week[weekday][class_number][room_number]
                if slot is None:
                    continue
                class_name = slot[0]
                instructor = slot[1]
                groups_list = slot[2]
                for group_name in groups_list:
                    result[(weekday, class_number, group_name)] = \
                        (room_number, class_name, instructor)
    return result


# week format is NOT parametrized
def empty_rooms(week, rooms):
    # result format: (weekday, class_number) -> [available rooms]
    result = dict()

    def to_string(x):
        if type(x) not in (str, int):
            return str(x.item())
        else:
            return str(x)

    def actual_occupied_room(room):
        if "online" in to_string(room):
            return False
        if not to_string(room).isnumeric():
            return False
        return not (week[weekday][class_number][int(to_string(room))] is None)

    for weekday in WEEKDAYS:
        result[weekday] = dict()
        for class_number in range(len(week[weekday])):
            # result[weekday][class_number] = list()
            occupied_rooms = set(filter(actual_occupied_room,
                                        map(to_string, week[weekday][class_number].keys())))
            available_rooms = rooms - occupied_rooms
            result[weekday][class_number] = sorted(available_rooms)
    return result


# week format: (weekday, class_number, group_name) -> (room_number, class_name, instructor)
def create_xlsx(block1_raw, block2_raw, groups, rooms):
    print('Started creating .xlsx')
    wb = Workbook()
    available_rooms_sheet = wb.active
    available_rooms_sheet.title = "Available rooms"
    total_block1 = wb.create_sheet("Block 1")
    total_block2 = wb.create_sheet("Block 2")
    rest_width_offset = 0

    write_available_rooms(available_rooms_sheet,
                          empty_rooms(block1_raw, rooms),
                          empty_rooms(block2_raw, rooms)
                         )

    block1 = parametrized(block1_raw)
    block2 = parametrized(block2_raw)

    YEARS = sorted(set(map(lambda x: x.split("-")[0], groups)), reverse=True)
    YEARS = sorted(YEARS, key=lambda x: x[0])

    for year in YEARS:
        year_groups = list(filter(lambda x: x.startswith(year), groups))
        entries_block1 = list(filter(lambda x: x[2] in year_groups, block1.keys()))
        entries_block2 = list(filter(lambda x: x[2] in year_groups, block2.keys()))
        year_sheet_block1 = wb.create_sheet(year + " (block 1)")
        year_sheet_block2 = wb.create_sheet(year + " (block 2)")

        def augmented(group_tag):
            return group_tag.replace("AI", "YAI").replace("DS", "YDS").replace("RO", "ZRO")

        # denoting groups and weekdays
        for column_number, group in enumerate(sorted(year_groups, key=augmented), 2):
            year_sheet_block1.cell(row=1,
                                   column=column_number,
                                   value=group)
            year_sheet_block2.cell(row=1,
                                   column=column_number,
                                   value=group)
            total_block1.cell(row=1,
                              column=column_number + rest_width_offset,
                              value=group)
            total_block2.cell(row=1,
                              column=column_number + rest_width_offset,
                              value=group)
        for weekday in WEEKDAYS:
            year_sheet_block1.cell(row=get_row_offset(weekday, 0),
                                   column=1,
                                   value=weekday)
            year_sheet_block2.cell(row=get_row_offset(weekday, 0),
                                   column=1,
                                   value=weekday)
            total_block1.cell(row=get_row_offset(weekday, 0),
                              column=1,
                              value=weekday)
            total_block2.cell(row=get_row_offset(weekday, 0),
                              column=1,
                              value=weekday)
            for class_number in range(0, 6):
                year_sheet_block1.cell(row=get_row_offset(weekday, class_number) + 1,
                                       column=1,
                                       value=PERIODS[class_number])
                year_sheet_block2.cell(row=get_row_offset(weekday, class_number) + 1,
                                       column=1,
                                       value=PERIODS[class_number])
                total_block1.cell(row=get_row_offset(weekday, class_number) + 1,
                                  column=1 + rest_width_offset,
                                  value=PERIODS[class_number])
                total_block2.cell(row=get_row_offset(weekday, class_number) + 1,
                                  column=1 + rest_width_offset,
                                  value=PERIODS[class_number])

        for slot in entries_block1:
            weekday = slot[0]
            class_number = slot[1]
            column_number = sorted(year_groups, key=augmented) \
                                .index(slot[2]) + 2
            row_number = get_row_offset(weekday, class_number)

            room_number = block1[slot][0]
            year_sheet_block1.cell(row=row_number + 3,
                                   column=column_number,
                                   value=room_number)
            total_block1.cell(row=row_number + 3,
                              column=column_number + rest_width_offset,
                              value=room_number)

            class_name = block1[slot][1]
            year_sheet_block1.cell(row=row_number + 1,
                                   column=column_number,
                                   value=class_name)
            total_block1.cell(row=row_number + 1,
                              column=column_number + rest_width_offset,
                              value=class_name)

            instructor = block1[slot][2]
            year_sheet_block1.cell(row=row_number + 2,
                                   column=column_number,
                                   value=instructor)
            total_block1.cell(row=row_number + 2,
                              column=column_number + rest_width_offset,
                              value=instructor)

        for slot in entries_block2:
            weekday = slot[0]
            class_number = slot[1]
            column_number = sorted(year_groups,
                                   key=augmented).index(slot[2]) + 2
            row_number = get_row_offset(weekday, class_number)

            room_number = block2[slot][0]
            year_sheet_block2.cell(row=row_number + 3,
                                   column=column_number,
                                   value=room_number)
            total_block2.cell(row=row_number + 3,
                              column=column_number + rest_width_offset,
                              value=room_number)

            class_name = block2[slot][1]
            year_sheet_block2.cell(row=row_number + 1,
                                   column=column_number,
                                   value=class_name)
            total_block2.cell(row=row_number + 1,
                              column=column_number + rest_width_offset,
                              value=class_name)

            instructor = block2[slot][2]
            year_sheet_block2.cell(row=row_number + 2,
                                   column=column_number,
                                   value=instructor)
            total_block2.cell(row=row_number + 2,
                              column=column_number + rest_width_offset,
                              value=instructor)

        rest_width_offset += len(year_groups) + 1

        prettify(total_block1,
                 rest_width_offset - 1,
                 rest_width_offset - len(year_groups) - 1)
        prettify(total_block2,
                 rest_width_offset - 1,
                 rest_width_offset - len(year_groups) - 1)
        prettify(year_sheet_block1,
                 len(year_groups))
        prettify(year_sheet_block2,
                 len(year_groups))

    wb.save(os.path.join("mysite/output_data", "schedule.xlsx"))
    # wb.save("schedule.xlsx")

def get_row_offset(weekday, class_number):
    weekday_number = WEEKDAYS.index(weekday)
    return 1 + weekday_number * 19 + 1 + class_number * 3


def write_available_rooms(sheet, available1, available2):
    font = Font(bold=True)
    border = Border(top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))
    fill = PatternFill(fill_type="solid",
                       start_color="22FF22",
                       end_color="22FF22")

    for weekday in range(len(WEEKDAYS)):
        sheet.cell(row=weekday * 7 + 1, column=1, value=WEEKDAYS[weekday])
        sheet.cell(row=weekday * 7 + 1, column=2, value="Rooms available in Block 1")
        sheet.cell(row=weekday * 7 + 1, column=3, value="Rooms available in Block 2")
        sheet[f"A{weekday * 7 + 1}"].font = font
        sheet[f"A{weekday * 7 + 1}"].border = border
        sheet[f"A{weekday * 7 + 1}"].fill = fill
        sheet[f"B{weekday * 7 + 1}"].border = border
        sheet[f"B{weekday * 7 + 1}"].fill = fill
        sheet[f"C{weekday * 7 + 1}"].border = border
        sheet[f"C{weekday * 7 + 1}"].fill = fill

        for class_number in range(0, 6):
            sheet.cell(row=weekday * 7 + class_number + 2, column=1, value=PERIODS[class_number])
            sheet.cell(row=weekday * 7 + class_number + 2, column=2,
                       value=" ".join(available1[WEEKDAYS[weekday]][class_number]))
            sheet.cell(row=weekday * 7 + class_number + 2, column=3,
                       value=" ".join(available2[WEEKDAYS[weekday]][class_number]))
        sheet.column_dimensions["A"].auto_size = True
        sheet.column_dimensions["B"].auto_size = True
        sheet.column_dimensions["C"].auto_size = True


def prettify(sheet, number_of_groups, class_period_column=0):
    # formatting the whole table
    alignment = Alignment(horizontal="center",
                          vertical="center",
                          wrap_text=True)
    for column in range(1 + number_of_groups):
        letter = get_column_letter(column + 1)
        sheet.column_dimensions[letter].auto_size = True
        sheet.column_dimensions[letter].width += 1
        for row in range(1, get_row_offset("Sun", 6) + 4):
            sheet[letter + str(row)].alignment = alignment

    # formatting group names
    font = Font(bold=True)
    border = Border(left=Side(border_style="thin", color='FF000000'),
                    right=Side(border_style="thin", color='FF000000'),
                    top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))
    fill = PatternFill(fill_type="darkGrid",
                       start_color="00AA00",
                       end_color="00AA00")
    for column in range(1 + number_of_groups):
        letter = get_column_letter(column + 1)
        sheet[letter + "1"].font = font
        sheet[letter + "1"].fill = fill
        sheet[letter + "1"].border = border

    # formatting weekday delimiters
    font = Font(bold=True)
    border = Border(top=Side(border_style="thin", color='FF000000'),
                    bottom=Side(border_style="thin", color='FF000000'))
    fill = PatternFill(fill_type="solid",
                       start_color="22FF22",
                       end_color="22FF22")
    for weekday in WEEKDAYS:
        row = str(get_row_offset(weekday, 0))
        for column in range(1 + number_of_groups):
            letter = get_column_letter(column + 1)
            sheet[letter + row].font = font
            sheet[letter + row].fill = fill
            sheet[letter + row].border = border

    # formatting class periods
    letter = get_column_letter(1 + class_period_column)
    sheet.column_dimensions[letter].width = 12
    border = Border(left=Side(border_style="thin",
                              color='FF000000'),
                    right=Side(border_style="thin",
                               color='FF000000'),
                    top=Side(border_style="thin",
                             color='FF000000'),
                    bottom=Side(border_style="thin",
                                color='FF000000'))
    fill = PatternFill(fill_type="darkGrid",
                       start_color="CCFFCC",
                       end_color="CCFFCC")
    for weekday in WEEKDAYS:
        for class_number in range(0, 6):
            start_merge_row = get_row_offset(weekday, class_number) + 1
            end_merge_row = start_merge_row + 2
            sheet.merge_cells(f"{letter}{start_merge_row}:{letter}{end_merge_row}")
            sheet[letter + str(start_merge_row)].border = border
            sheet[letter + str(start_merge_row)].fill = fill
            row_offset = get_row_offset(weekday, class_number)
            for column in range(2, 2 + number_of_groups):
                sheet.cell(column=column, row=row_offset + 1).border = \
                    Border(left=Side(border_style="thin", color='FF000000'),
                           right=Side(border_style="thin", color='FF000000'),
                           top=Side(border_style="thin", color='FF000000'))
                sheet.cell(column=column, row=row_offset + 2).border = \
                    Border(left=Side(border_style="thin", color='FF000000'),
                           right=Side(border_style="thin", color='FF000000'))
                sheet.cell(column=column, row=row_offset + 3).border = \
                    Border(left=Side(border_style="thin", color='FF000000'),
                           right=Side(border_style="thin", color='FF000000'),
                           bottom=Side(border_style="thin", color='FF000000'))

    # merging inner cells with same subjects
    for row_number in range(1, get_row_offset("Sun", 6) + 4):
        current_value = ""
        start_column = 0
        for column in range(2, 2 + number_of_groups):
            if column == 2:
                start_column = 2
                current_value = sheet.cell(row=row_number,
                                           column=column).value
                current_value = str(current_value).strip()
                continue
            if (current_value == str(sheet.cell(row=row_number,
                                                column=column).value).strip() and
                    (current_value != "None")):
                if column == number_of_groups + 1:
                    sheet.merge_cells(start_column=start_column,
                                      end_column=column,
                                      start_row=row_number,
                                      end_row=row_number)
                continue
            elif (current_value ==
                  str(sheet.cell(row=row_number, column=column).value).strip() and
                  (current_value == "None")):
                start_column = column
            else:
                if start_column != column - 1:
                    sheet.merge_cells(start_column=start_column,
                                      end_column=column - 1,
                                      start_row=row_number,
                                      end_row=row_number)
                start_column = column
                current_value = sheet.cell(row=row_number,
                                           column=column).value
                current_value = str(current_value).strip()

    # "online X" -> "ONLINE"
    for row_number in range(1, get_row_offset("Sun", 6) + 4):
        for column in range(2, 2 + number_of_groups):
            current_value = str(sheet.cell(row=row_number, column=column).value)
            if current_value.startswith("online"):
                sheet.cell(row=row_number, column=column, value="ONLINE")